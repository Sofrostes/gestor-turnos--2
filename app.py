import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile
import os
import shutil

# ------------------------------------------------------------
# 1. CARGA EXACTA DEL EXCEL (CELDA POR CELDA)
# ------------------------------------------------------------
@st.cache_data
def load_excel_data(file_path, sheet_name="MAYO 2026"):
    """Carga el Excel leyendo celda por celda y guarda las referencias de celda"""
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        
        # Convertir toda la hoja a una matriz de valores
        all_data = []
        for row in sheet.iter_rows(values_only=True):
            all_data.append([cell if cell is not None else "" for cell in row])
        
        if not all_data:
            return None, None, "No se encontraron datos"
        
        # Buscar la fila de encabezados
        header_row_idx = None
        for i, row in enumerate(all_data):
            if row and any("NOMBRE" in str(cell).upper() or "AGENTE" in str(cell).upper() for cell in row):
                header_row_idx = i
                break
        
        if header_row_idx is None:
            return None, None, "No se encontró la fila de encabezados"
        
        # Identificar columnas importantes
        zona_col = 0      # Columna A
        codigo_col = 2    # Columna C
        nombre_col = 4    # Columna E
        primer_dia_col = 5  # Columna F
        
        # Guardar referencias de celdas para poder modificarlas después
        referencias_celdas = {}  # (fila, columna) -> celda de openpyxl
        # También guardar la posición de cada agente
        agentes = []
        start_row = header_row_idx + 1
        
        # Primero, obtener todas las filas de agentes y sus referencias
        for row_idx in range(start_row, len(all_data)):
            row = all_data[row_idx]
            if not row or len(row) < primer_dia_col + 1:
                continue
            
            zona = str(row[zona_col]).strip() if zona_col < len(row) else ""
            codigo = str(row[codigo_col]).strip() if codigo_col < len(row) else ""
            nombre = str(row[nombre_col]).strip() if nombre_col < len(row) else ""
            
            # Filtrar agentes válidos
            if not codigo or codigo == "0" or codigo == "":
                continue
            if not nombre or nombre == "0" or nombre == "":
                continue
            if "DESPLAZADO" in nombre.upper() or "VACANTE" in nombre.upper():
                continue
            
            # Extraer turnos y guardar referencias de celdas
            turnos = []
            celdas_turnos = []
            for dia_idx in range(primer_dia_col, min(primer_dia_col + 31, len(row))):
                valor = row[dia_idx]
                turnos.append(str(valor).strip() if valor is not None and valor != "" else "")
                # Guardar referencia de la celda (fila, columna)
                celdas_turnos.append((row_idx, dia_idx))
            
            # Asegurar 31 días
            while len(turnos) < 31:
                turnos.append("")
            
            agente = {
                "zona": zona,
                "codigo": codigo,
                "nombre": nombre,
                "turnos": turnos,
                "fila": row_idx,
                "col_inicio": primer_dia_col,
                "celdas_turnos": celdas_turnos
            }
            agentes.append(agente)
            
            # Guardar referencia de la celda de código y nombre (no se modifican)
            referencias_celdas[(row_idx, codigo_col)] = sheet.cell(row=row_idx+1, column=codigo_col+1)
            referencias_celdas[(row_idx, nombre_col)] = sheet.cell(row=row_idx+1, column=nombre_col+1)
        
        if not agentes:
            return None, None, "No se encontraron agentes válidos"
        
        # Organizar por zona
        agentes_por_zona = {}
        for ag in agentes:
            zona = ag["zona"]
            if zona not in agentes_por_zona:
                agentes_por_zona[zona] = []
            agentes_por_zona[zona].append(ag)
        
        return agentes_por_zona, wb, sheet, f"✅ Cargados {len(agentes)} agentes en {len(agentes_por_zona)} zonas"
    
    except Exception as e:
        return None, None, None, f"❌ Error: {str(e)}"

# ------------------------------------------------------------
# 2. FUNCIONES PARA EDITAR Y GUARDAR TURNOS
# ------------------------------------------------------------
def intercambiar_turnos_en_memoria(agentes_por_zona, zona, agente1_idx, agente2_idx, dia):
    """Intercambia los turnos entre dos agentes de la misma zona en memoria"""
    if zona not in agentes_por_zona:
        return False
    agentes = agentes_por_zona[zona]
    if agente1_idx >= len(agentes) or agente2_idx >= len(agentes):
        return False
    
    turno1 = agentes[agente1_idx]["turnos"][dia]
    turno2 = agentes[agente2_idx]["turnos"][dia]
    
    agentes[agente1_idx]["turnos"][dia] = turno2
    agentes[agente2_idx]["turnos"][dia] = turno1
    
    return True

def guardar_cambios_en_excel(agentes_por_zona, wb, sheet, archivo_ruta):
    """Guarda los turnos modificados en el archivo Excel original"""
    try:
        # Recorrer todos los agentes y actualizar las celdas correspondientes
        for zona, agentes in agentes_por_zona.items():
            for agente in agentes:
                for dia_idx, turno in enumerate(agente["turnos"]):
                    if dia_idx < len(agente["celdas_turnos"]):
                        fila, col = agente["celdas_turnos"][dia_idx]
                        # Actualizar la celda en el objeto sheet
                        celda = sheet.cell(row=fila+1, column=col+1)
                        celda.value = turno if turno else None
        
        # Guardar el workbook
        wb.save(archivo_ruta)
        return True, "✅ Cambios guardados permanentemente en el archivo Excel"
    except Exception as e:
        return False, f"❌ Error al guardar: {str(e)}"

# ------------------------------------------------------------
# 3. INTERFAZ DE STREAMLIT
# ------------------------------------------------------------
st.set_page_config(page_title="Cuadrantes Metrovalencia", layout="wide", page_icon="📅")

st.title("📅 Cuadrante de Servicios - Metrovalencia")
st.markdown("### Visualización y edición de turnos (con guardado en Excel)")

# Inicializar estado de sesión
if 'agentes_por_zona' not in st.session_state:
    st.session_state.agentes_por_zona = None
if 'wb' not in st.session_state:
    st.session_state.wb = None
if 'sheet' not in st.session_state:
    st.session_state.sheet = None
if 'archivo_path' not in st.session_state:
    st.session_state.archivo_path = None
if 'archivo_cargado' not in st.session_state:
    st.session_state.archivo_cargado = False

# Sidebar
with st.sidebar:
    st.header("📁 Cargar archivo")
    
    # Opción: trabajar con copia o con original
    modo_guardado = st.radio(
        "Modo de guardado",
        ["Trabajar con copia (seguro)", "Guardar en original (sobrescribe)"],
        help="Con copia no se modifica el archivo original. Con original los cambios son permanentes."
    )
    
    uploaded_file = st.file_uploader("Selecciona el Excel", type=["xlsx"])
    
    if uploaded_file is not None:
        # Guardar archivo temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        
        if modo_guardado == "Trabajar con copia (seguro)":
            # Crear una copia para trabajar
            copia_path = tmp_path.replace(".xlsx", "_copia.xlsx")
            shutil.copy2(tmp_path, copia_path)
            archivo_trabajo = copia_path
            st.info("📋 Trabajando con una copia del archivo. Los cambios no afectarán al original.")
        else:
            archivo_trabajo = tmp_path
            st.warning("⚠️ Los cambios se guardarán DIRECTAMENTE en el archivo original")
        
        agentes_por_zona, wb, sheet, msg = load_excel_data(archivo_trabajo, "MAYO 2026")
        
        if agentes_por_zona:
            st.success(msg)
            st.session_state.agentes_por_zona = agentes_por_zona
            st.session_state.wb = wb
            st.session_state.sheet = sheet
            st.session_state.archivo_path = archivo_trabajo
            st.session_state.archivo_cargado = True
        else:
            st.error(msg)
    
    # Selector de zona (solo si hay datos)
    if st.session_state.archivo_cargado and st.session_state.agentes_por_zona:
        zonas = list(st.session_state.agentes_por_zona.keys())
        zona_seleccionada = st.selectbox("📍 Seleccionar zona", zonas, key="zona_select")
        st.session_state.zona_actual = zona_seleccionada
        
        # Mostrar estadísticas
        agentes_zona = st.session_state.agentes_por_zona[zona_seleccionada]
        st.metric("👥 Agentes en zona", len(agentes_zona))
        
        st.markdown("---")
        st.markdown("### ✏️ Edición")
        modo_edicion = st.checkbox("Activar modo edición", value=False, key="modo_edicion")
        
        if modo_edicion:
            st.info("💡 Selecciona dos agentes y un día para intercambiar turnos")

# Contenido principal
if not st.session_state.archivo_cargado:
    st.info("👈 Por favor, carga el archivo Excel en el panel lateral")
    with st.expander("📖 Instrucciones"):
        st.markdown("""
        ### Formato esperado:
        - El archivo debe contener la hoja **MAYO 2026**
        - Los agentes deben tener:
          - **Columna A**: Identificador de zona (JC, AV, AA, FO, etc.)
          - **Columna C**: Código del agente (CF)
          - **Columna E**: Nombre del agente
          - **Columna F en adelante**: Turnos del día 1 al 31
        
        ### Funcionalidades:
        - Visualización exacta de los turnos (celda por celda)
        - Filtro automático: solo agentes con CF y nombre
        - **Intercambio de turnos** entre agentes de la misma zona
        - **Guardado permanente** en el archivo Excel
        """)
else:
    zona_actual = st.session_state.get('zona_actual', list(st.session_state.agentes_por_zona.keys())[0])
    agentes = st.session_state.agentes_por_zona[zona_actual]
    
    # Preparar datos para la tabla
    dias = list(range(1, 32))
    
    # Crear DataFrame para visualización
    df_data = []
    for idx, agente in enumerate(agentes):
        row = {
            "IDX": idx,
            "Código": agente["codigo"],
            "Agente": agente["nombre"]
        }
        for i, dia in enumerate(dias):
            turno = agente["turnos"][i] if i < len(agente["turnos"]) else ""
            # Resaltar turnos que son números (servicios especiales)
            if turno and turno.isdigit():
                row[f"D{i+1}"] = f"🔹 {turno}"
            else:
                row[f"D{i+1}"] = turno if turno else "—"
        df_data.append(row)
    
    df = pd.DataFrame(df_data)
    
    # Mostrar tabla principal
    st.markdown(f"## 📊 Zona {zona_actual}")
    
    # Configurar columnas para la tabla
    column_config = {
        "IDX": st.column_config.NumberColumn("ID", width="small"),
        "Código": st.column_config.TextColumn("Código", width="small"),
        "Agente": st.column_config.TextColumn("Agente", width="medium"),
    }
    for i in range(1, 32):
        column_config[f"D{i}"] = st.column_config.TextColumn(f"Día {i}", width="small")
    
    st.dataframe(
        df,
        use_container_width=True,
        height=min(600, len(agentes) * 35 + 38),
        column_config=column_config,
        hide_index=True
    )
    
    # Modo edición: interfaz para intercambiar turnos
    if st.session_state.get('modo_edicion', False):
        st.markdown("---")
        st.markdown("## 🔄 Intercambiar turnos")
        st.warning("⚠️ Solo se pueden intercambiar turnos entre agentes de la misma zona")
        
        col1, col2, col3 = st.columns([2, 2, 1])
        
        with col1:
            agentes_nombres = [f"{a['nombre']} (Cód: {a['codigo']})" for a in agentes]
            agente1 = st.selectbox("Agente origen", agentes_nombres, key="agente1")
            agente1_idx = agentes_nombres.index(agente1)
        
        with col2:
            agente2 = st.selectbox("Agente destino", agentes_nombres, key="agente2")
            agente2_idx = agentes_nombres.index(agente2)
        
        with col3:
            dia_seleccionado = st.selectbox("Día", list(range(1, 32)), key="dia_intercambio")
            dia_idx = dia_seleccionado - 1
        
        # Mostrar turnos actuales
        turno1 = agentes[agente1_idx]["turnos"][dia_idx] if dia_idx < len(agentes[agente1_idx]["turnos"]) else ""
        turno2 = agentes[agente2_idx]["turnos"][dia_idx] if dia_idx < len(agentes[agente2_idx]["turnos"]) else ""
        
        st.info(f"**Turno actual:** {agentes[agente1_idx]['nombre']} → {turno1 if turno1 else '—'} | {agentes[agente2_idx]['nombre']} → {turno2 if turno2 else '—'}")
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            if st.button("🔄 Intercambiar (solo memoria)", use_container_width=True):
                if intercambiar_turnos_en_memoria(st.session_state.agentes_por_zona, zona_actual, agente1_idx, agente2_idx, dia_idx):
                    st.success("✅ Turnos intercambiados en memoria")
                    st.rerun()
                else:
                    st.error("❌ Error al intercambiar")
        
        with col_btn2:
            if st.button("💾 Intercambiar Y GUARDAR en Excel", type="primary", use_container_width=True):
                # Primero intercambiar en memoria
                if intercambiar_turnos_en_memoria(st.session_state.agentes_por_zona, zona_actual, agente1_idx, agente2_idx, dia_idx):
                    # Luego guardar en el archivo Excel
                    ok, msg = guardar_cambios_en_excel(
                        st.session_state.agentes_por_zona,
                        st.session_state.wb,
                        st.session_state.sheet,
                        st.session_state.archivo_path
                    )
                    if ok:
                        st.success(msg)
                        st.balloons()
                        st.rerun()
                    else:
                        st.error(msg)
                else:
                    st.error("❌ Error al intercambiar")
        
        # Botón para guardar todos los cambios pendientes
        st.markdown("---")
        if st.button("💾 Guardar TODOS los cambios en Excel", type="secondary", use_container_width=True):
            ok, msg = guardar_cambios_en_excel(
                st.session_state.agentes_por_zona,
                st.session_state.wb,
                st.session_state.sheet,
                st.session_state.archivo_path
            )
            if ok:
                st.success(msg)
                st.balloons()
            else:
                st.error(msg)

# Footer
st.markdown("---")
st.markdown("📅 **Metrovalencia - Gestión de Cuadrantes** | Lectura exacta de Excel | Edición con guardado permanente")